import requests
import openpyxl
import os
import tkinter as tk
from tkinter import messagebox

def search_card():
    card_name = entry_card_name.get()
    
    # Search for the card on Scryfall
    url = f"https://api.scryfall.com/cards/named?fuzzy={card_name}"
    response = requests.get(url)
    data = response.json()

    # Extracting relevant information
    name = data.get('name', '')
    colors = ', '.join(data.get('colors', [])) if data.get('colors') else ''
    card_url = data.get('scryfall_uri', '')

    # Check if the Excel file already exists
    file_name = "MyCardLibrary.xlsx"
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Card Data"
        ws.append(["Card Name", "Colors", "URL", "Count"])
    else:
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active

    # Check if the card already exists in the file
    card_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        if row[0].value == name:
            card_count = ws.cell(row=row[0].row, column=4).value
            ws.cell(row=row[0].row, column=4, value=card_count + 1)
            break
    else:
        ws.append([name, colors, card_url, 1])

    # Save the Excel file
    wb.save(file_name)
    messagebox.showinfo("Success", f"Data saved to {file_name}")

# Create GUI
root = tk.Tk()
root.title("Card Search and Save")

label_card_name = tk.Label(root, text="Enter the card name:")
label_card_name.grid(row=0, column=0, padx=10, pady=10)

entry_card_name = tk.Entry(root)
entry_card_name.grid(row=0, column=1, padx=10, pady=10)

btn_search = tk.Button(root, text="Search and Save", command=search_card)
btn_search.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()
